from io import BytesIO

from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.viewsets import ModelViewSet

from django.http import HttpResponse
from django.utils.translation import gettext as _

from auth_tenant.mixins import TenantQuerySetMixin
from auth_tenant.permissions import page_action_permission

from .models import Sale
from .serializers import SaleCreateSerializer, SaleDetailSerializer, SaleListSerializer


class SaleViewSet(TenantQuerySetMixin, ModelViewSet):
    queryset = Sale.objects.select_related("client", "created_by").all()
    serializer_class = SaleListSerializer
    http_method_names = ["get", "post", "head", "options"]

    filterset_fields = ["payment_type"]
    ordering_fields = ["created_at", "total_amount", "client__name", "created_by__first_name"]
    ordering = ["-created_at"]

    def get_permissions(self):
        if self.action == "retrieve":
            return [page_action_permission("sales", "detail")()]
        if self.action in ("list", "export_excel"):
            return [page_action_permission("sales", "read")()]
        return [page_action_permission("sales", "write")()]

    def get_serializer_class(self):
        if self.action == "create":
            return SaleCreateSerializer
        if self.action == "retrieve":
            return SaleDetailSerializer
        return SaleListSerializer

    def get_queryset(self):
        qs = super().get_queryset()

        payment_type = self.request.query_params.get("payment_type")
        if payment_type in {
            Sale.PaymentType.CASH,
            Sale.PaymentType.CARD,
            Sale.PaymentType.TRANSFER,
            Sale.PaymentType.DEBT,
        }:
            qs = qs.filter(payment_type=payment_type)

        client_ids_raw = self.request.query_params.get("client_ids")
        if client_ids_raw:
            ids = [i.strip() for i in client_ids_raw.split(",") if i.strip()]
            if ids:
                qs = qs.filter(client__id__in=ids)

        seller_ids_raw = self.request.query_params.get("seller_ids")
        if seller_ids_raw:
            ids = [i.strip() for i in seller_ids_raw.split(",") if i.strip()]
            if ids:
                qs = qs.filter(created_by__id__in=ids)

        date_from = self.request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        amount_from = self.request.query_params.get("amount_from")
        if amount_from:
            try:
                qs = qs.filter(total_amount__gte=amount_from)
            except (ValueError, TypeError):
                pass

        amount_to = self.request.query_params.get("amount_to")
        if amount_to:
            try:
                qs = qs.filter(total_amount__lte=amount_to)
            except (ValueError, TypeError):
                pass

        if self.action == "retrieve":
            qs = qs.prefetch_related("items__product")
        return qs

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ValidationError(
                {"detail": _("Excel export requires openpyxl package on backend.")}
            ) from exc

        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        headers = ["ID", "Client", "Seller", "Total", "Payment", "Date"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for sale in queryset.iterator(chunk_size=500):
            client_name = sale.client.name if sale.client else ""
            seller_name = ""
            if sale.created_by:
                parts = [sale.created_by.first_name or "", getattr(sale.created_by, "last_name", "") or ""]
                seller_name = " ".join(p for p in parts if p)
            ws.append([
                str(sale.id),
                client_name,
                seller_name,
                float(sale.total_amount),
                sale.get_payment_type_display(),
                sale.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="sales_export.xlsx"'
        return response
