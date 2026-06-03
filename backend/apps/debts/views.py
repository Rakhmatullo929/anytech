from decimal import Decimal
from io import BytesIO

from django.db import transaction
from django.db.models import (
    Count,
    DecimalField,
    Exists,
    ExpressionWrapper,
    F,
    OuterRef,
    Q,
    Subquery,
    Sum,
    Max,
)
from django.http import Http404, HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.mixins import ListModelMixin, RetrieveModelMixin
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet

from apps.auth_tenant.mixins import TenantQuerySetMixin
from apps.auth_tenant.permissions import page_action_permission

from .models import Debt, Payment
from .serializers import (
    CustomerDebtSummarySerializer,
    DebtDetailSerializer,
    DebtListSerializer,
    DebtPaymentListSerializer,
    PaymentCreateSerializer,
)


class DebtViewSet(TenantQuerySetMixin, ListModelMixin, RetrieveModelMixin, GenericViewSet):
    queryset = Debt.objects.select_related("client").all()
    serializer_class = DebtListSerializer

    filterset_fields = ["status"]
    ordering_fields = ["created_at", "total_amount", "paid_amount", "remaining_amount", "status", "deadline", "client__name"]
    ordering = ["-created_at"]

    def get_permissions(self):
        if self.action == "retrieve":
            return [page_action_permission("debts", "detail")()]
        if self.action in ("list", "export_excel"):
            return [page_action_permission("debts", "read")()]
        return [page_action_permission("debts", "write")()]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return DebtDetailSerializer
        if self.action == "pay":
            return PaymentCreateSerializer
        return DebtListSerializer

    def get_queryset(self):
        qs = super().get_queryset().annotate(
            remaining_amount=ExpressionWrapper(
                F("total_amount") - F("paid_amount"),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            )
        )

        status = self.request.query_params.get("status")
        if status in {Debt.Status.ACTIVE, Debt.Status.CLOSED}:
            qs = qs.filter(status=status)

        client_ids_raw = self.request.query_params.get("client_ids")
        if client_ids_raw:
            ids = [i.strip() for i in client_ids_raw.split(",") if i.strip()]
            if ids:
                qs = qs.filter(client__id__in=ids)

        date_from = self.request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        deadline_from = self.request.query_params.get("deadline_from")
        if deadline_from:
            qs = qs.filter(deadline__gte=deadline_from)

        deadline_to = self.request.query_params.get("deadline_to")
        if deadline_to:
            qs = qs.filter(deadline__lte=deadline_to)

        amount_from = self.request.query_params.get("amount_from")
        if amount_from:
            try:
                qs = qs.filter(total_amount__gte=amount_from)
            except (ValueError, TypeError):
                pass

        amount_to = self.request.query_params.get("amount_to")
        if amount_to:
            try:
                qs = qs.filter(total_amount__lte=amount_to)
            except (ValueError, TypeError):
                pass

        if self.action == "retrieve":
            qs = qs.prefetch_related("payments")
        return qs

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ValidationError(
                {"detail": _("Excel export requires openpyxl package on backend.")}
            ) from exc

        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Debts"

        headers = ["ID", "Client", "Total", "Paid", "Remaining", "Status", "Deadline", "Created"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for debt in queryset.iterator(chunk_size=500):
            ws.append([
                str(debt.id),
                debt.client.name if debt.client else "",
                float(debt.total_amount),
                float(debt.paid_amount),
                float(debt.remaining),
                debt.get_status_display(),
                debt.deadline.strftime("%Y-%m-%d") if debt.deadline else "",
                debt.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="debts_export.xlsx"'
        return response

    @action(detail=True, methods=["post"], url_path="pay")
    def pay(self, request, pk=None):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        amount = serializer.validated_data["amount"]

        with transaction.atomic():
            try:
                debt = (
                    self.get_queryset()
                    .select_for_update()
                    .select_related("client")
                    .get(pk=pk)
                )
            except Debt.DoesNotExist:
                raise Http404

            if debt.status == Debt.Status.CLOSED:
                raise ValidationError({"detail": _("Debt is already closed.")})

            if amount > debt.remaining:
                raise ValidationError(
                    {"detail": _("Amount exceeds remaining debt of %(remaining)s.") % {
                        "remaining": debt.remaining
                    }}
                )

            payment_method = serializer.validated_data.get(
                "payment_method", Payment.PaymentMethod.CASH
            )
            Payment.objects.create(
                debt=debt,
                amount=amount,
                payment_method=payment_method,
                cashier=request.user,
            )

            debt.paid_amount += amount
            if debt.paid_amount >= debt.total_amount:
                debt.status = Debt.Status.CLOSED
            debt.save(update_fields=["paid_amount", "status", "updated_at"])

        debt = (
            self.get_queryset()
            .select_related("client")
            .prefetch_related("payments")
            .get(pk=pk)
        )
        return Response(
            DebtDetailSerializer(debt, context=self.get_serializer_context()).data
        )


class DebtPaymentViewSet(ListModelMixin, GenericViewSet):
    queryset = Payment.objects.select_related("debt__client", "cashier").all()
    serializer_class = DebtPaymentListSerializer
    ordering_fields = ["created_at", "amount"]
    ordering = ["-created_at"]

    def get_permissions(self):
        return [page_action_permission("debts", "read")()]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated or not user.tenant_id:
            return Payment.objects.none()

        qs = super().get_queryset().filter(debt__tenant_id=user.tenant_id)

        customer_id = self.request.query_params.get("customer_id")
        if customer_id:
            qs = qs.filter(debt__client_id=customer_id)

        payment_method = self.request.query_params.get("payment_method")
        if payment_method in {m[0] for m in Payment.PaymentMethod.choices}:
            qs = qs.filter(payment_method=payment_method)

        cashier_ids_raw = self.request.query_params.get("cashier_ids")
        if cashier_ids_raw:
            ids = [i.strip() for i in cashier_ids_raw.split(",") if i.strip()]
            if ids:
                qs = qs.filter(cashier_id__in=ids)

        date_from = self.request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        return qs

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ValidationError(
                {"detail": _("Excel export requires openpyxl package on backend.")}
            ) from exc

        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Debt Payments"

        headers = ["ID", "Client", "Amount", "Payment Method", "Cashier", "Date"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for payment in queryset.iterator(chunk_size=500):
            client = payment.debt.client
            cashier = payment.cashier
            cashier_name = (
                " ".join(filter(None, [cashier.first_name, cashier.last_name]))
                if cashier
                else ""
            )
            ws.append([
                str(payment.id),
                client.name if client else "",
                float(payment.amount),
                payment.get_payment_method_display(),
                cashier_name,
                payment.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="debt_payments_export.xlsx"'
        return response


class CustomerDebtSummaryViewSet(ListModelMixin, GenericViewSet):
    """Aggregated debt balances per customer (only customers with remaining debt > 0)."""

    serializer_class = CustomerDebtSummarySerializer

    def get_permissions(self):
        return [page_action_permission("debts", "read")()]

    def _build_queryset(self):
        user = self.request.user
        if not user.is_authenticated or not user.tenant_id:
            return Debt.objects.none().values("client")

        today = timezone.now().date()

        has_overdue = Debt.objects.filter(
            tenant_id=user.tenant_id,
            client=OuterRef("client"),
            status=Debt.Status.ACTIVE,
            deadline__isnull=False,
            deadline__lt=today,
        )

        last_payment_subq = (
            Payment.objects.filter(
                debt__client=OuterRef("client"),
                debt__tenant_id=user.tenant_id,
            )
            .order_by("-created_at")
            .values("created_at")[:1]
        )

        base = Debt.objects.filter(tenant_id=user.tenant_id)

        search = self.request.query_params.get("search", "").strip()
        if search:
            base = base.filter(client__name__icontains=search)

        qs = (
            base.values("client")
            .annotate(
                client_name=F("client__name"),
                client_phone=F("client__phone"),
                total_debt=Sum("total_amount"),
                total_paid=Sum("paid_amount"),
                remaining=ExpressionWrapper(
                    Sum("total_amount") - Sum("paid_amount"),
                    output_field=DecimalField(max_digits=14, decimal_places=2),
                ),
                debt_count=Count("id"),
                last_debt_date=Max("created_at"),
                last_payment_date=Subquery(last_payment_subq),
                is_overdue=Exists(has_overdue),
            )
            .filter(remaining__gt=0)
        )

        date_from = self.request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(last_debt_date__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(last_debt_date__date__lte=date_to)

        amount_from = self.request.query_params.get("amount_from")
        if amount_from:
            try:
                qs = qs.filter(remaining__gte=amount_from)
            except (ValueError, TypeError):
                pass

        amount_to = self.request.query_params.get("amount_to")
        if amount_to:
            try:
                qs = qs.filter(remaining__lte=amount_to)
            except (ValueError, TypeError):
                pass

        ordering = self.request.query_params.get("ordering", "-last_debt_date")
        allowed = frozenset({
            "client_name", "-client_name",
            "total_debt", "-total_debt",
            "total_paid", "-total_paid",
            "remaining", "-remaining",
            "debt_count", "-debt_count",
            "last_debt_date", "-last_debt_date",
        })
        qs = qs.order_by(ordering if ordering in allowed else "-last_debt_date")

        return qs

    def get_queryset(self):
        return self._build_queryset()

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        user = request.user
        if not user.is_authenticated or not user.tenant_id:
            return Response({
                "totalOutstanding": "0.00",
                "customersWithDebt": 0,
                "overdueCustomers": 0,
                "averageDebt": "0.00",
            })

        today = timezone.now().date()
        base = Debt.objects.filter(tenant_id=user.tenant_id)

        per_client = (
            base.values("client")
            .annotate(
                client_remaining=ExpressionWrapper(
                    Sum("total_amount") - Sum("paid_amount"),
                    output_field=DecimalField(max_digits=14, decimal_places=2),
                )
            )
            .filter(client_remaining__gt=0)
        )

        customers_count = per_client.count()
        total_outstanding = (
            per_client.aggregate(total=Sum("client_remaining"))["total"] or Decimal("0")
        )
        avg_debt = total_outstanding / customers_count if customers_count else Decimal("0")

        overdue_count = (
            base.filter(
                status=Debt.Status.ACTIVE,
                deadline__isnull=False,
                deadline__lt=today,
            )
            .values("client")
            .distinct()
            .count()
        )

        return Response({
            "totalOutstanding": str(total_outstanding),
            "customersWithDebt": customers_count,
            "overdueCustomers": overdue_count,
            "averageDebt": str(round(avg_debt, 2)),
        })

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ValidationError(
                {"detail": _("Excel export requires openpyxl package on backend.")}
            ) from exc

        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Debts"

        headers = ["Client", "Phone", "Total Debt", "Total Paid", "Remaining", "Debt Count", "Last Debt Date", "Last Payment Date", "Status"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        serializer = CustomerDebtSummarySerializer(many=True)
        for row_data in queryset.iterator(chunk_size=500):
            status = serializer.child.get_status(row_data)
            ws.append([
                row_data.get("client_name", ""),
                row_data.get("client_phone", ""),
                float(row_data.get("total_debt") or 0),
                float(row_data.get("total_paid") or 0),
                float(row_data.get("remaining") or 0),
                row_data.get("debt_count", 0),
                row_data["last_debt_date"].strftime("%Y-%m-%d %H:%M") if row_data.get("last_debt_date") else "",
                row_data["last_payment_date"].strftime("%Y-%m-%d %H:%M") if row_data.get("last_payment_date") else "",
                status,
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="customer_debts_export.xlsx"'
        return response
