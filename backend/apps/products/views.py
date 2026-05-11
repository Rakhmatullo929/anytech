from decimal import Decimal, InvalidOperation
from io import BytesIO

from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from django.db import transaction
from django.db.models import Avg, DecimalField, F, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils.translation import gettext as _

from auth_tenant.mixins import TenantQuerySetMixin
from auth_tenant.permissions import page_action_permission

from .models import Category, Product, ProductPurchase

# ---------------------------------------------------------------------------
# Excel import helpers (module-level so they're not recreated per request)
# ---------------------------------------------------------------------------

_EXCEL_HEADER_NAMES = {"name", "название", "nomi", "nom"}


def _excel_clean(value) -> str:
    return str(value).strip() if value is not None else ""


def _excel_parse_row(row) -> dict:
    def _get(i):
        return _excel_clean(row[i]) if i < len(row) else ""

    return {
        "name": _get(0),
        "sku": _get(1),
        "category": _get(2),
        "stock": _get(3),
        "purchasePrice": _get(4),
    }


from .serializers import (
    CategoryBulkDeleteSerializer,
    CategorySerializer,
    ProductBulkCreateExcelSerializer,
    ProductBulkDeleteSerializer,
    ProductDetailSerializer,
    ProductListSerializer,
    ProductPurchaseBulkDeleteSerializer,
    ProductPurchaseSerializer,
    ProductSerializer,
    ProductUpdateSerializer,
)


class ProductViewSet(TenantQuerySetMixin, ModelViewSet):
    queryset = (
        Product.objects.all()
        .prefetch_related("images")
        .annotate(
            total_quantity=Coalesce(Sum("purchases__quantity"), Value(0)),
            total_purchase_amount=Coalesce(
                Sum(
                    F("purchases__quantity") * F("purchases__unit_price"),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
                Value(0),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
            average_purchase_price=Coalesce(
                Avg("purchases__unit_price"),
                Value(0),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            ),
        )
    )
    serializer_class = ProductSerializer
    http_method_names = ["get", "post", "put", "patch", "delete", "head", "options"]

    search_fields = ["name", "sku"]
    ordering_fields = ["name", "created_at", "total_quantity"]
    ordering = ["-created_at"]

    def get_queryset(self):
        # available_quantity = total_quantity: purchases are physically deducted on sale (FIFO).
        queryset = super().get_queryset().annotate(
            available_quantity=F("total_quantity")
        )
        category_id = self.request.query_params.get("category_id")
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        in_stock = self.request.query_params.get("in_stock")
        if in_stock == "true":
            queryset = queryset.filter(total_quantity__gt=0)
        return queryset

    def get_permissions(self):
        if self.action in ("list", "search"):
            return [page_action_permission("products", "read")()]
        if self.action == "retrieve":
            return [page_action_permission("products", "detail")()]
        return [page_action_permission("products", "write")()]

    def get_serializer_class(self):
        if self.action in ("list", "search"):
            return ProductListSerializer
        if self.action == "retrieve":
            return ProductDetailSerializer
        if self.action in ("update", "partial_update"):
            return ProductUpdateSerializer
        return ProductSerializer

    @action(detail=False, methods=["get"], url_path="search")
    def search(self, request):
        return self.list(request)

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        serializer = ProductBulkDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        deleted_count, _details = self.get_queryset().filter(id__in=ids).delete()
        return Response({"deleted": deleted_count})

    @action(detail=False, methods=["get"], url_path="download-excel-template")
    def download_excel_template(self, request):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ValidationError(
                {"detail": _("Excel export requires openpyxl package on backend.")}
            ) from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = ["name", "sku", "category", "stock", "purchasePrice"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.append(["Example Product", "SKU-001", "Electronics", 10, "9.99"])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products_template.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="bulk-create-excel")
    def bulk_create_excel(self, request):
        serializer = ProductBulkCreateExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        excel_file = serializer.validated_data["file"]

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValidationError(
                {"file": _("Excel import requires openpyxl package on backend.")}
            ) from exc

        try:
            # read_only=True uses a streaming SAX parser — much lower memory for large files
            workbook = load_workbook(filename=excel_file, data_only=True, read_only=True)
        except Exception as exc:
            raise ValidationError({"file": _("Invalid Excel file.")}) from exc

        sheet = workbook.active or workbook.worksheets[0]
        raw_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        if not raw_rows:
            raise ValidationError({"file": _("Excel file is empty.")})

        parsed = [_excel_parse_row(r) for r in raw_rows]

        if parsed and parsed[0]["name"].lower() in _EXCEL_HEADER_NAMES:
            parsed = parsed[1:]

        # (1-indexed, +2 because header was row 1)
        data_rows = [(i + 2, r) for i, r in enumerate(parsed) if r["name"] or r["sku"]]

        if not data_rows:
            raise ValidationError({"file": _("No product rows found in Excel.")})

        tenant = request.user.tenant

        # Pre-fetch existing SKUs to detect conflicts in one query
        file_skus = [r["sku"] for _, r in data_rows if r["sku"]]
        existing_skus = set(
            Product.objects.filter(tenant=tenant, sku__in=file_skus).values_list("sku", flat=True)
        )

        row_errors: list[dict] = []
        valid_rows: list[dict] = []
        seen_skus: set[str] = set()

        for row_num, row in data_rows:
            errors: list[str] = []

            if not row["name"]:
                errors.append(_("Name is required."))

            sku = row["sku"] or None
            if sku:
                if sku in existing_skus:
                    errors.append(_("SKU '%(sku)s' already exists.") % {"sku": sku})
                elif sku in seen_skus:
                    errors.append(_("Duplicate SKU '%(sku)s' in file.") % {"sku": sku})
                else:
                    seen_skus.add(sku)

            stock: int | None = None
            if row["stock"]:
                try:
                    stock = int(row["stock"])
                    if stock < 0:
                        errors.append(_("Stock must be a non-negative integer."))
                        stock = None
                except (ValueError, TypeError):
                    errors.append(_("Stock must be an integer."))

            purchase_price: Decimal | None = None
            if row["purchasePrice"]:
                try:
                    purchase_price = Decimal(row["purchasePrice"])
                    if purchase_price < 0:
                        errors.append(_("Purchase price must be non-negative."))
                        purchase_price = None
                except InvalidOperation:
                    errors.append(_("Purchase price must be a number."))

            if errors:
                row_errors.append({"row": row_num, "errors": errors})
            else:
                valid_rows.append({
                    "name": row["name"],
                    "sku": sku,
                    "category": row["category"],
                    "stock": stock,
                    "purchasePrice": purchase_price,
                })

        if not valid_rows:
            return Response(
                {"created": 0, "errors": row_errors, "results": []},
                status=status.HTTP_200_OK,
            )

        # --- Category resolution (case-insensitive, auto-create) ---
        all_existing_categories = list(Category.objects.filter(tenant=tenant))
        category_map: dict[str, Category] = {c.name.lower(): c for c in all_existing_categories}

        new_cats_by_lower: dict[str, str] = {}
        for row in valid_rows:
            cat_name = row["category"]  # already stripped by _excel_clean
            if cat_name:
                lower = cat_name.lower()
                if lower not in category_map and lower not in new_cats_by_lower:
                    new_cats_by_lower[lower] = cat_name

        if new_cats_by_lower:
            created_cats = Category.objects.bulk_create(
                [Category(tenant=tenant, name=name) for name in new_cats_by_lower.values()],
                ignore_conflicts=True,
            )
            fetched: dict[str, Category] = {c.name.lower(): c for c in created_cats if c.pk}
            # Fetch any that were skipped due to concurrent inserts (rare race condition)
            conflict_names = [name for lower, name in new_cats_by_lower.items() if lower not in fetched]
            if conflict_names:
                fetched.update(
                    {c.name.lower(): c for c in Category.objects.filter(tenant=tenant, name__in=conflict_names)}
                )
            category_map.update(fetched)

        # --- Bulk create products and optional purchases ---
        with transaction.atomic():
            products_to_create = [
                Product(
                    tenant=tenant,
                    name=row["name"],
                    sku=row["sku"],
                    category=category_map.get(row["category"].lower()) if row["category"] else None,
                )
                for row in valid_rows
            ]
            created_products = Product.objects.bulk_create(products_to_create)

            purchases_to_create = []
            for product, row in zip(created_products, valid_rows):
                stock = row["stock"]
                if stock is not None and stock > 0:
                    purchases_to_create.append(
                        ProductPurchase(
                            product=product,
                            quantity=stock,
                            unit_price=row["purchasePrice"] if row["purchasePrice"] is not None else Decimal("0.00"),
                        )
                    )

            if purchases_to_create:
                ProductPurchase.objects.bulk_create(purchases_to_create)

        created_ids = [p.id for p in created_products]
        result_qs = self.get_queryset().filter(id__in=created_ids)
        response_data = ProductListSerializer(result_qs, many=True, context={"request": request}).data

        return Response(
            {"created": len(response_data), "errors": row_errors, "results": response_data},
            status=status.HTTP_200_OK,
        )


class CategoryViewSet(TenantQuerySetMixin, ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    http_method_names = ["get", "post", "put", "delete", "head", "options"]

    search_fields = ["name"]
    ordering_fields = ["name", "created_at"]
    ordering = ["-created_at"]

    def get_permissions(self):
        if self.action == "list":
            return [page_action_permission("categories", "read")()]
        if self.action == "retrieve":
            return [page_action_permission("categories", "detail")()]
        return [page_action_permission("categories", "write")()]

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        serializer = CategoryBulkDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        deleted_count, _details = self.get_queryset().filter(id__in=ids).delete()
        return Response({"deleted": deleted_count})


class ProductPurchaseViewSet(ModelViewSet):
    queryset = ProductPurchase.objects.all().select_related("product")
    serializer_class = ProductPurchaseSerializer
    http_method_names = ["get", "post", "put", "delete", "head", "options"]

    search_fields = ["product__name"]
    ordering_fields = ["created_at", "quantity", "unit_price"]
    ordering = ["-created_at"]

    def get_queryset(self):
        queryset = self.queryset.filter(product__tenant_id=self.request.user.tenant_id)
        product_id = self.request.query_params.get("product_id")
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        return queryset

    def get_permissions(self):
        if self.action == "list":
            return [page_action_permission("products", "read")()]
        if self.action == "retrieve":
            return [page_action_permission("products", "detail")()]
        return [page_action_permission("products", "write")()]

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        serializer = ProductPurchaseBulkDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        deleted_count, _details = self.get_queryset().filter(id__in=ids).delete()
        return Response({"deleted": deleted_count})
