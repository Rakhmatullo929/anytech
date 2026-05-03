from rest_framework import status
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from django.db import transaction
from django.db.models import Count
from django.utils.translation import gettext as _

from auth_tenant.mixins import TenantQuerySetMixin
from auth_tenant.permissions import page_action_permission

from .models import Client, Group
from .serializers import (
    ClientBulkDeleteSerializer,
    ClientBulkCreateExcelSerializer,
    ClientDetailSerializer,
    ClientSerializer,
    GroupAddClientsSerializer,
    GroupBulkDeleteSerializer,
    GroupDetailSerializer,
    GroupListSerializer,
)


class ClientViewSet(TenantQuerySetMixin, ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    http_method_names = ["get", "post", "put", "patch", "delete", "head", "options"]

    search_fields = ["name", "last_name", "middle_name", "phone"]
    ordering_fields = ["name", "created_at"]
    ordering = ["-created_at"]

    def get_permissions(self):
        if self.action in ("list", "search"):
            return [page_action_permission("clients", "read")()]
        if self.action == "retrieve":
            return [page_action_permission("clients", "detail")()]
        return [page_action_permission("clients", "write")()]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ClientDetailSerializer
        return ClientSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        group_id = self.request.query_params.get("group_id")
        if group_id:
            qs = qs.filter(groups__id=group_id)
        if self.action == "retrieve":
            qs = qs.prefetch_related(
                "sales__items__product",
                "sales__debt",
                "debts",
                "groups",
            )
        return qs

    @action(detail=False, methods=["get"], url_path="search")
    def search(self, request):
        return self.list(request)

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        serializer = ClientBulkDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        deleted_count, _details = self.get_queryset().filter(id__in=ids).delete()
        return Response({"deleted": deleted_count}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="bulk-create-excel")
    def bulk_create_excel(self, request):
        serializer = ClientBulkCreateExcelSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        excel_file = serializer.validated_data["file"]

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValidationError(
                {"file": _("Excel import requires openpyxl package on backend.")}
            ) from exc

        try:
            workbook = load_workbook(filename=excel_file, data_only=True)
        except Exception as exc:
            raise ValidationError({"file": _("Invalid Excel file.")}) from exc

        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            raise ValidationError({"file": _("Excel file is empty.")})

        def _clean(value):
            if value is None:
                return ""
            return str(value).strip()

        rows = [(_clean(row[0]) if len(row) > 0 else "", _clean(row[1]) if len(row) > 1 else "") for row in raw_rows]

        # Skip header row if it looks like a header.
        if rows:
            first_name, first_phone = rows[0]
            header_like = first_name.lower() in {"name", "имя", "mijoz", "client"} and first_phone.lower() in {
                "phone",
                "телефон",
                "telefon",
            }
            if header_like:
                rows = rows[1:]

        valid_rows = [(name, phone) for name, phone in rows if name or phone]
        if not valid_rows:
            raise ValidationError({"file": _("No client rows found in Excel.")})

        errors = []
        phones_in_file = []
        for idx, (name, phone) in enumerate(valid_rows, start=1):
            if not name:
                errors.append(_("Row %(idx)d: name is required.") % {"idx": idx})
            if not phone:
                errors.append(_("Row %(idx)d: phone is required.") % {"idx": idx})
            if name and phone:
                phones_in_file.append(phone)

        duplicate_phones_in_file = {phone for phone in phones_in_file if phones_in_file.count(phone) > 1}
        if duplicate_phones_in_file:
            errors.append(
                _("Duplicate phones in file: %(phones)s.")
                % {"phones": ", ".join(sorted(duplicate_phones_in_file))}
            )

        tenant = request.user.tenant
        existing_phones = set(
            Client.objects.filter(tenant=tenant, phone__in=phones_in_file).values_list("phone", flat=True)
        )
        if existing_phones:
            errors.append(
                _("Phones already exist: %(phones)s.")
                % {"phones": ", ".join(sorted(existing_phones))}
            )

        if errors:
            raise ValidationError({"file": errors})

        with transaction.atomic():
            clients_to_create = [
                Client(tenant=tenant, name=name, phone=phone, phones=[phone])
                for name, phone in valid_rows
            ]
            created_clients = Client.objects.bulk_create(clients_to_create)

        response_data = ClientSerializer(created_clients, many=True, context={"request": request}).data
        return Response({"created": len(response_data), "results": response_data}, status=status.HTTP_201_CREATED)


class GroupViewSet(TenantQuerySetMixin, ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupListSerializer
    http_method_names = ["get", "post", "put", "patch", "delete", "head", "options"]

    search_fields = ["name"]
    ordering_fields = ["name", "created_at"]
    ordering = ["-created_at"]

    def get_permissions(self):
        if self.action == "retrieve":
            return [page_action_permission("groups", "detail")()]
        if self.action in ("list", "search"):
            return [page_action_permission("groups", "read")()]
        return [page_action_permission("groups", "write")()]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return GroupDetailSerializer
        return GroupListSerializer

    def get_queryset(self):
        return super().get_queryset().annotate(clients_count=Count("clients", distinct=True))

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        serializer = GroupBulkDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        deleted_count, _details = self.get_queryset().filter(id__in=ids).delete()
        return Response({"deleted": deleted_count}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="add-clients")
    def add_clients(self, request, pk=None):
        group = self.get_object()
        serializer = GroupAddClientsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        clients = Client.objects.filter(tenant=request.user.tenant, id__in=ids)
        group.clients.add(*clients)
        return Response(status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="remove-clients")
    def remove_clients(self, request, pk=None):
        group = self.get_object()
        serializer = GroupAddClientsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data["ids"]
        clients = Client.objects.filter(tenant=request.user.tenant, id__in=ids)
        group.clients.remove(*clients)
        return Response(status=status.HTTP_200_OK)
